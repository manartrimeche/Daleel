"""
Export service — generate downloadable files from compliance data.

Supports:
  • Excel (.xlsx) via openpyxl
  • CSV fallback (no extra dependency)

Exportable datasets:
  • Exigences d'un document (obligations, sanctions, conditions, interdictions)
  • Feuille de route de conformité (roadmap)
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any

from app.database import get_collection
from app.services.criticality_service import compute_criticality_score, score_to_level

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
# Exigence export
# ─────────────────────────────────────────────────────────────

_TYPE_LABELS = {
    "obligation": "Obligation",
    "sanction": "Sanction",
    "condition": "Condition",
    "prohibition": "Interdiction",
}


async def export_exigences_file(
    db: Any,
    document_id: str,
    *,
    format: str = "xlsx",
) -> tuple[bytes, str, str]:
    """
    Export all exigences for a document as Excel or CSV.

    For each exigence, attempts to join related actions and criticality
    scores when available. Returns (content_bytes, media_type, filename).
    """
    # Fetch document metadata
    doc = await get_collection("documents").find_one({"id": document_id})
    doc_name = (doc.get("filename", document_id) if doc else document_id)
    safe_name = "".join(
        c if c.isalnum() or c in " _-" else "_" for c in doc_name
    ).strip().rstrip("_")
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d")

    # Fetch exigences sorted by page then type
    exigences = await (
        get_collection("exigences")
        .find({"document_id": document_id})
        .sort([("page_number", 1), ("exigence_type", 1)])
        .to_list(length=10_000)
    )

    # Fetch related actions (keyed by exigence_id)
    exigence_ids = [e["id"] for e in exigences if e.get("id")]
    actions_by_exigence: dict[str, list[dict]] = {}
    if exigence_ids:
        actions_cursor = get_collection("actions").find(
            {"exigence_id": {"$in": exigence_ids}}
        )
        async for act in actions_cursor:
            eid = act.get("exigence_id")
            actions_by_exigence.setdefault(eid, []).append(act)

    # Build sanctions context per article for inherited boost
    sanctions_by_article: dict[str, str] = {}
    for e in exigences:
        if e.get("exigence_type") == "sanction":
            art = e.get("article_reference") or ""
            prev = sanctions_by_article.get(art, "")
            sanctions_by_article[art] = (prev + " " + (e.get("text") or "")).strip()

    # Build rows
    headers = [
        "Page",
        "Article",
        "Type",
        "Exigence",
        "Confiance",
        "Action concrète",
        "Preuve requise",
        "Conditions",
        "Criticité",
        "Score",
    ]

    rows: list[list] = []
    for e in exigences:
        eid = e.get("id", "")
        art = e.get("article_reference") or ""
        etype = e.get("exigence_type", "")
        actions = actions_by_exigence.get(eid, [])
        sanctions_ctx = sanctions_by_article.get(art, "")

        if actions:
            for act in actions:
                score, _ = compute_criticality_score(
                    act, sanctions_context=sanctions_ctx
                )
                level = score_to_level(score)
                rows.append([
                    e.get("page_number", ""),
                    art,
                    _TYPE_LABELS.get(etype, etype),
                    e.get("text", ""),
                    e.get("confidence_score", ""),
                    act.get("action_precise", ""),
                    act.get("preuve") or "",
                    " | ".join(act.get("conditions") or []),
                    level.capitalize(),
                    round(score, 2),
                ])
        else:
            # No action linked — still export the exigence with inline scoring
            fake_action = {"modalite": etype, "action_precise": e.get("text", "")}
            score, _ = compute_criticality_score(
                fake_action, sanctions_context=sanctions_ctx
            )
            level = score_to_level(score)
            rows.append([
                e.get("page_number", ""),
                art,
                _TYPE_LABELS.get(etype, etype),
                e.get("text", ""),
                e.get("confidence_score", ""),
                "",
                "",
                "",
                level.capitalize(),
                round(score, 2),
            ])

    # Build summary stats
    summary = {
        "document": doc_name,
        "total": len(exigences),
        "by_type": {},
        "by_level": {"Critique": 0, "Importante": 0, "Secondaire": 0},
    }
    for e in exigences:
        t = _TYPE_LABELS.get(e.get("exigence_type", ""), "Autre")
        summary["by_type"][t] = summary["by_type"].get(t, 0) + 1
    for row in rows:
        lvl = row[8] if len(row) > 8 else ""
        if lvl in summary["by_level"]:
            summary["by_level"][lvl] += 1

    if format == "xlsx":
        return _exigences_to_xlsx(headers, rows, safe_name, timestamp, summary)
    return _exigences_to_csv(headers, rows, safe_name, timestamp)


def _exigences_to_csv(
    headers: list[str],
    rows: list[list],
    safe_name: str,
    timestamp: str,
) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    content = buf.getvalue().encode("utf-8-sig")
    filename = f"exigences_{safe_name}_{timestamp}.csv"
    return content, "text/csv; charset=utf-8", filename


def _exigences_to_xlsx(
    headers: list[str],
    rows: list[list],
    safe_name: str,
    timestamp: str,
    summary: dict,
) -> tuple[bytes, str, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV")
        return _exigences_to_csv(headers, rows, safe_name, timestamp)

    wb = Workbook()

    # ── Sheet 1 : Résumé ──
    ws_sum = wb.active
    ws_sum.title = "Résumé"
    title_font = Font(bold=True, size=14, color="1F4E79")
    sub_font = Font(bold=True, size=11)
    ws_sum.append(["Analyse des exigences réglementaires"])
    ws_sum["A1"].font = title_font
    ws_sum.append([])
    ws_sum.append(["Document", summary.get("document", "")])
    ws_sum["A3"].font = sub_font
    ws_sum.append(["Total exigences", summary.get("total", 0)])
    ws_sum["A4"].font = sub_font
    ws_sum.append(["Date d'export", timestamp])
    ws_sum["A5"].font = sub_font
    ws_sum.append([])

    ws_sum.append(["Répartition par type"])
    ws_sum[f"A{ws_sum.max_row}"].font = sub_font
    for type_name, count in summary.get("by_type", {}).items():
        ws_sum.append(["", type_name, count])

    ws_sum.append([])
    ws_sum.append(["Répartition par criticité"])
    ws_sum[f"A{ws_sum.max_row}"].font = sub_font

    level_colors = {
        "Critique": "FFC7CE",
        "Importante": "FFEB9C",
        "Secondaire": "C6EFCE",
    }
    for level_name, count in summary.get("by_level", {}).items():
        row_num = ws_sum.max_row + 1
        ws_sum.append(["", level_name, count])
        fill_color = level_colors.get(level_name)
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws_sum.cell(row=row_num, column=2).fill = fill
            ws_sum.cell(row=row_num, column=3).fill = fill

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.column_dimensions["C"].width = 12

    # ── Sheet 2 : Exigences détaillées ──
    ws = wb.create_sheet("Exigences")

    # Header row
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="B0B0B0"),
    )
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows with criticality coloring
    level_fills = {
        "Critique": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Importante": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Secondaire": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }
    level_fonts = {
        "Critique": Font(color="9C0006"),
        "Importante": Font(color="9C6500"),
        "Secondaire": Font(color="006100"),
    }

    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_idx, row_data in enumerate(rows, 2):
        ws.append(row_data)
        level = row_data[8] if len(row_data) > 8 else ""

        # Apply colored fill to the entire row
        if level in level_fills:
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = level_fills[level]
                cell.border = thin_border

            # Bold criticality cell
            crit_cell = ws.cell(row=row_idx, column=9)
            crit_cell.font = Font(bold=True, color=level_fonts[level].color)

        # Wrap long text columns
        for col_idx in (4, 6, 7, 8):
            ws.cell(row=row_idx, column=col_idx).alignment = wrap_align

    # Auto-width (capped)
    col_widths = {
        1: 8,    # Page
        2: 14,   # Article
        3: 14,   # Type
        4: 55,   # Exigence
        5: 12,   # Confiance
        6: 45,   # Action
        7: 30,   # Preuve
        8: 30,   # Conditions
        9: 14,   # Criticité
        10: 8,   # Score
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Freeze header row + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    filename = f"exigences_{safe_name}_{timestamp}.xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return content, media_type, filename


async def export_roadmap_file(
    db: Any,
    profile_id: str,
    *,
    format: str = "xlsx",
    organization_id: str | None = None,
) -> tuple[bytes, str, str]:
    """
    Generate an export file for the compliance roadmap.

    Returns (content_bytes, media_type, filename).
    """
    from app.services.roadmap_service import generate_roadmap

    roadmap = await generate_roadmap(db, profile_id, organization_id=organization_id)
    profile_name = roadmap.get("profile_name", profile_id)[:40]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    plan = roadmap.get("ordered_plan", [])

    headers = [
        "Position",
        "Action ID",
        "Loi",
        "Article",
        "Modalité",
        "Action précise",
        "Conditions",
        "Preuve",
        "Criticité",
        "Score",
        "Dépendances",
    ]

    rows = []
    for item in plan:
        rows.append([
            item.get("position", ""),
            item.get("action_id", ""),
            item.get("loi_code", ""),
            item.get("article_key", ""),
            item.get("modalite", ""),
            item.get("action_precise", ""),
            " | ".join(item.get("conditions") or []),
            item.get("preuve", ""),
            item.get("criticality_level", ""),
            item.get("criticality_score", ""),
            ", ".join(item.get("depends_on_ids") or []),
        ])

    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in profile_name).strip()

    if format == "xlsx":
        return _to_xlsx(headers, rows, safe_name, timestamp, roadmap)
    else:
        return _to_csv(headers, rows, safe_name, timestamp)


def _to_csv(
    headers: list[str],
    rows: list[list],
    safe_name: str,
    timestamp: str,
) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    filename = f"roadmap_{safe_name}_{timestamp}.csv"
    return content, "text/csv; charset=utf-8", filename


def _to_xlsx(
    headers: list[str],
    rows: list[list],
    safe_name: str,
    timestamp: str,
    roadmap: dict,
) -> tuple[bytes, str, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV export")
        return _to_csv(headers, rows, safe_name, timestamp)

    wb = Workbook()

    # ── Summary sheet ──
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    ws_summary.append(["Feuille de route de conformité"])
    ws_summary.append(["Profil", roadmap.get("profile_name", "")])
    ws_summary.append(["Total actions", roadmap.get("total_actions", 0)])
    by_level = roadmap.get("by_level", {})
    ws_summary.append(["Critique", by_level.get("critique", 0)])
    ws_summary.append(["Importante", by_level.get("importante", 0)])
    ws_summary.append(["Secondaire", by_level.get("secondaire", 0)])
    gen = roadmap.get("generated_at")
    ws_summary.append(["Généré le", str(gen) if gen else timestamp])

    title_font = Font(bold=True, size=14)
    ws_summary["A1"].font = title_font

    # ── Plan sheet ──
    ws_plan = wb.create_sheet("Plan d'action")
    ws_plan.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws_plan.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    level_colors = {
        "critique": "FFC7CE",
        "importante": "FFEB9C",
        "secondaire": "C6EFCE",
    }

    for row_idx, row in enumerate(rows, 2):
        ws_plan.append(row)
        level = row[8] if len(row) > 8 else ""
        if level in level_colors:
            fill = PatternFill(start_color=level_colors[level], end_color=level_colors[level], fill_type="solid")
            for col_idx in range(1, len(row) + 1):
                ws_plan.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-width columns
    for col in ws_plan.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception as exc:
                logger.debug("Cell width calc skipped: %s", exc)
        ws_plan.column_dimensions[col_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    filename = f"roadmap_{safe_name}_{timestamp}.xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return content, media_type, filename
